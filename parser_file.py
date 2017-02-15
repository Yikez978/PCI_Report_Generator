from people import People
import sys
import datetime
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
import xlwings as xw
import openpyxl

def load_main_file(filename, main_dic):

    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    line_number = len(sheet["A"])
    print "Filename: ", filename, "Lines: ", line_number

    racker_name_list = []
    contractor_name_list = []
      
    for row in range(9, line_number+1):
        combi_name = sheet['A'+str(row)].value.encode('ascii','ignore').lower().strip()
        status = sheet['C'+str(row)].value.encode('ascii','ignore').lower().strip()
        company = sheet['J'+str(row)].value.encode('ascii','ignore').lower().strip()
        time = sheet['D'+str(row)].value
        active_status = sheet['K'+str(row)].value.encode('ascii','ignore').lower().strip()

        if status != "completed":
            status = "incompleted"
        if company != "contractor":
            company = "rackspace"

        if time > datetime.datetime(2017, 1, 1) and active_status == "active":
            if combi_name not in main_dic:
                people = People(combi_name)
                people.status = status
                people.company = company

                main_dic[combi_name] = people

            #name duplicate
            else:
                if status == "completed":
                    main_dic[combi_name].status = status

            
            last_name, first_name = combi_name.split(",")
            last_name = last_name.strip()
            first_name = first_name.strip()
            if company == "rackspace" and (first_name, last_name) not in racker_name_list:
                racker_name_list.append((first_name, last_name))
            if company != "rackspace" and (first_name, last_name) not in contractor_name_list:
                contractor_name_list.append((first_name, last_name))

    print "Active racker registered in 2017: ", len(main_dic.keys())

def generate_name_list(name, dic):
    wb = openpyxl.Workbook()
    name_list = []

    for key in dic.keys():
        if dic[key].company == name:
            name_list.append(key)

    if len(name_list) == 0:
        print "Something wrong with list"

    ws = wb.get_sheet_by_name('Sheet')

    for index in range(len(name_list)):
        ws.cell(row=index+1, column=1).value = name_list[index]
 
    wb.save(name+".xlsx")

def update_people(filename, dic):

    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    line_number = len(sheet["A"])
    print "Filename: ", filename, "Lines: ", line_number

    name_not_in_main_dic = []
 
    for row in range(1, line_number+1):
        first_name = sheet['A'+str(row)].value.encode('ascii','ignore').lower().strip()
        last_name = sheet['B'+str(row)].value.encode('ascii','ignore').lower().strip()

        combi_name = last_name+', '+first_name

        bu = sheet['D'+str(row)].value
        if bu:
            bu = bu.encode('ascii','ignore').lower().strip()

        if combi_name not in dic:
            name_not_in_main_dic.append(combi_name)
        else:
            dic[combi_name].bu = bu 

    return name_not_in_main_dic

def find_missing_contractor(filename, dic):

    wb = openpyxl.load_workbook(filename)
    sheets = wb.sheetnames

    missing_name_list = []

    for i in range(len(sheets)):
        print "On sheet name : ", sheets[i]
        sheet = wb.get_sheet_by_name(wb.sheetnames[i])
        line_number = len(sheet["A"])
        print "Filename: ", filename, "lines: ", line_number
 
        for row in range(2, line_number+1):
            first_name = sheet['A'+str(row)].value.encode('ascii','ignore').lower().strip()
            last_name = sheet['B'+str(row)].value.encode('ascii','ignore').lower().strip()
            sso = sheet['C'+str(row)].value
            email = sheet['D'+str(row)].value
            bu1 = sheet['E'+str(row)].value
            bu2 = sheet['F'+str(row)].value
            bu3 = sheet['G'+str(row)].value
            bu4 = sheet['H'+str(row)].value

           # if " [c]" in combi_name:
           #     combi_name = combi_name.replace(" [c]", "").strip()

           # last_name = combi_name.split(" ")[-1]
           # first_name = combi_name[:-len(last_name)].strip()

            combi_name = last_name + ", " + first_name 
            
            if combi_name not in dic.keys():
                missing_name_list.append([first_name, last_name, sso, email, bu1, bu2, bu3, bu4])
                print combi_name

    w = xw.Book()
    xw.Range('A1').value = missing_name_list 

def find_missing_racker(filename, dic):

    wb = openpyxl.load_workbook(filename)
    sheets = wb.sheetnames

    missing_name_list = []

    for i in range(len(sheets)):
        print "On sheet name : ", sheets[i]
        sheet = wb.get_sheet_by_name(wb.sheetnames[i])
        line_number = len(sheet["A"])
        print "first_name: ", filename, "lines: ", line_number
 
        for row in range(2, line_number+1):
            combi_name = sheet['A'+str(row)].value.encode('ascii','ignore').lower().strip()
            
            if combi_name not in dic:
                missing_name_list.append([combi_name])
    print missing_name_list

    w = xw.Book()
    xw.Range('A1').value = missing_name_list 

def generate_report(dic):
    print "Generate Report...Pci_Report.xlsx"
    racker_bu_dic = {}
    contractor_bu_dic = {}

    racker_without_bu_numlist = []
    contractor_without_bu_numlist = []

    for key in dic.keys():
        if dic[key].company == "rackspace":
            bu = dic[key].bu
            if bu:
                #racker_bu_dic[key][0, 0] complete number, total
                if bu not in racker_bu_dic.keys():
                    racker_bu_dic[bu] = [0, 0, [], []]
                if dic[key].status == "completed":
                    racker_bu_dic[bu][0] += 1
                    racker_bu_dic[bu][2].append(key)
                else:
                    racker_bu_dic[bu][3].append(key)
                racker_bu_dic[bu][1] +=1

            else:
                racker_without_bu_numlist.append(dic[key].combi_name)
        else:
            bu = dic[key].bu
            if bu:
                if bu not in contractor_bu_dic.keys():
                    contractor_bu_dic[bu] = [0, 0, [], []]
                if dic[key].status == "completed":
                    contractor_bu_dic[bu][0] += 1
                    contractor_bu_dic[bu][2].append(key)
                else: 
                    contractor_bu_dic[bu][3].append(key)
                contractor_bu_dic[bu][1] +=1

            else:
                contractor_without_bu_numlist.append(dic[key].combi_name)


    racker_bu_name_list = []
    contractor_bu_name_list = []
    
    for key in racker_bu_dic:
        if key not in racker_bu_name_list:
            racker_bu_name_list.append(key)
        
    for key in contractor_bu_dic:
        if key not in contractor_bu_name_list:
            contractor_bu_name_list.append(key)


    all_bu_name_list = list(set(contractor_bu_name_list)|set(racker_bu_name_list))

    wb = openpyxl.Workbook()

    #generate report analysis
    analysis_output(wb, all_bu_name_list, racker_bu_name_list, contractor_bu_name_list, racker_bu_dic, contractor_bu_dic)

    #generate multiple sheets by superviser name
    generate_sheets_by_supervisor(wb, all_bu_name_list, dic, racker_bu_dic, contractor_bu_dic)

    remove_sheet_name = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(remove_sheet_name)

    wb.save("PCI_Report.xlsx")

def analysis_output(wb, all_bu_name_list, racker_bu_name_list, contractor_bu_name_list, racker_bu_dic, contractor_bu_dic):

    r_bu_len = len(racker_bu_name_list)
    c_bu_len = len(contractor_bu_name_list)
    result_list = []

    racker_completed_total = 0
    racker_registered_total = 0
    racker_completed_ratio = 0

    contractor_completed_total = 0
    contractor_registered_total = 0
    contractor_completed_ratio = 0

    for bu_name in all_bu_name_list:
        racker_completed = 0
        racker_all  = 0

        contractor_complated = 0
        contractor_all = 0

        racker_percentage = 0
        contractor_percentrage = 0

        if bu_name in racker_bu_dic.keys():
            racker_completed, racker_all, a, b = racker_bu_dic[bu_name]
            racker_completed_total += racker_completed
            racker_registered_total += racker_all
            racker_percentage = str(racker_completed*100/racker_all) + " %"
            
        if bu_name in contractor_bu_dic.keys():
            contractor_complated, contractor_all, a, b = contractor_bu_dic[bu_name]
            contractor_completed_total += contractor_complated
            contractor_registered_total += contractor_all
            contractor_percentrage = str(contractor_complated*100/contractor_all) + " %"

        r_c_completed = racker_completed + contractor_complated
        r_c_registerd = racker_all + contractor_all
        r_c_ratio = str(r_c_completed*100/r_c_registerd) + " %"

        result_list.append([bu_name, racker_completed, racker_all, racker_percentage, contractor_complated, contractor_all, contractor_percentrage, r_c_completed, r_c_registerd, r_c_ratio] )


    racker_completed_ratio = str(racker_completed_total*100/racker_registered_total) + " %"
    contractor_completed_ratio = str(contractor_completed_total*100/contractor_registered_total) + " %"

    r_c_completed_tatal = racker_completed_total + contractor_completed_total
    r_c_registered_total = racker_registered_total + contractor_registered_total
    r_c_ratio_total = str(r_c_completed_tatal*100/r_c_registered_total) + " %"


    result_list.insert(0, ["Total", racker_completed_total, racker_registered_total, racker_completed_ratio, contractor_completed_total, contractor_registered_total, contractor_completed_ratio, r_c_completed_tatal, r_c_registered_total, r_c_ratio_total ])

    ws = wb.create_sheet("Overall")
    ws.cell(row=2, column=1).value = "Racker Completed"
    ws.cell(row=3, column=1).value = "Racker Registered"
    ws.cell(row=4, column=1).value = "Racker Ratio"

    ws.cell(row=5, column=1).value = "Contractor Completed"
    ws.cell(row=6, column=1).value = "Contractor Registered"
    ws.cell(row=7, column=1).value = "Contractor Ratio"

    ws.cell(row=8, column=1).value = "R & C Completed"
    ws.cell(row=9, column=1).value = "R & C Registered"
    ws.cell(row=10, column=1).value = "R & C Ratio"

    ws.cell(row=2, column=1).font = Font(bold=True)
    ws.cell(row=2, column=1).font = Font(bold=True)
    ws.cell(row=3, column=1).font = Font(bold=True)
    ws.cell(row=4, column=1).font = Font(bold=True)
    ws.cell(row=5, column=1).font = Font(bold=True)
    ws.cell(row=6, column=1).font = Font(bold=True)
    ws.cell(row=7, column=1).font = Font(bold=True)
    ws.cell(row=8, column=1).font = Font(bold=True)
    ws.cell(row=9, column=1).font = Font(bold=True)
    ws.cell(row=10, column=1).font = Font(bold=True)
    ft = Font(color=colors.RED)

    start = "A"
    for index in range(len(result_list)):
        ws.column_dimensions[start].width = 20
        start = chr(ord(start)+1)
        for item in range(len(result_list[0])):
            name = result_list[index][item]
            if item+1 == 1:
                ws.cell(row=item+1, column=index+2).font = Font(bold=True)
                name = upper_first_letter(name)
            ws.cell(row=item+1, column=index+2).value = name
            # given color
            # if 0 < int(result_list[index][3].split(" ")[0].strip()) < 30:
            #     ws.cell(row=item+1, column=index+2).font = Font(color=colors.RED) 

def upper_first_letter(name):
    name_list = name.split()
    result_list = []
    for i in name_list:
        i = i[0].upper() + i[1:]
        result_list.append(i)
    return " ".join(result_list)

def generate_sheets_by_supervisor(wb, all_bu_name_list, dic, racker_bu_dic, contractor_bu_dic):

    for key in all_bu_name_list:
        racker_completed = []
        racker_incompleted = []
        contractor_complated = []
        contractor_incomplated = []

        if key in racker_bu_dic.keys():
            racker_completed = racker_bu_dic[key][2]
            racker_incompleted = racker_bu_dic[key][3]
        if key in contractor_bu_dic.keys():
            contractor_complated = contractor_bu_dic[key][2]
            contractor_incomplated = contractor_bu_dic[key][3]

        write_to_new_sheet(wb, key, racker_completed, racker_incompleted, contractor_complated, contractor_incomplated)


def write_to_new_sheet(wb, superviser, l1, l2, l3, l4):

    index = 1
    superviser = upper_first_letter(superviser)
    ws = wb.create_sheet(superviser)
    ws.cell(row=1, column=1).value = "Racker Completed"
    ws.cell(row=1, column=2).value = "Racker InCompleted"
    ws.cell(row=1, column=3).value = "Contractor Completed"
    ws.cell(row=1, column=4).value = "Contractor InCompleted"

    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2).font = Font(bold=True)
    ws.cell(row=1, column=3).font = Font(bold=True)
    ws.cell(row=1, column=4).font = Font(bold=True)

    ws.column_dimensions["A"].width = max([len(i) for i in l1] + [20])
    ws.column_dimensions["B"].width = max([len(i) for i in l2] + [20])
    ws.column_dimensions["C"].width = max([len(i) for i in l3] + [20])
    ws.column_dimensions["D"].width = max([len(i) for i in l4] + [20])


    for l in [l1, l2, l3, l4]:
        for r in range(2, len(l)+2):
            name = upper_first_letter(l[r-2])
            ws.cell(row=r, column=index).value = name
        index += 1

def compare_with_main_and_Contractor_not_regitst_list(dic):

    wb = openpyxl.load_workbook("TES_FULL_CONTRACTOR_LIST.xlsx")
    sheets = wb.sheetnames
    sheet = wb.get_sheet_by_name(sheets[0])
    line_number = len(sheet["A"])
    print "lines: ", line_number

    missing_name_list = []

    for row in range(2, line_number+1):
        combi_name = sheet['E'+str(row)].value.encode('ascii','ignore').lower().strip()
        
        last_name = combi_name.split(" ")[-1].strip()
        first_name = combi_name[:-len(last_name)].strip()

        combi_name = last_name + ", " + first_name
        
        if combi_name not in dic:
            missing_name_list.append(combi_name)
    print missing_name_list
    print "length : ", len(missing_name_list)


#compare with Contractor_not_regiestered_list.xlsx 
    wb = openpyxl.load_workbook("Contractor_not_regi_list.xlsx")
    sheets = wb.sheetnames
    sheet = wb.get_sheet_by_name(sheets[0])
    line_number = len(sheet["A"])
    print "lines: ", line_number

    for row in range(1, line_number+1):
        first_name = sheet['A'+str(row)].value.encode('ascii','ignore').lower().strip()
        last_name = sheet['B'+str(row)].value.encode('ascii','ignore').lower().strip()
        
        combi_name = last_name + ", " + first_name
        
        if combi_name in missing_name_list:
            print combi_name, "in the list"
            missing_name_list.remove(combi_name)

    print missing_name_list
    print "length : ", len(missing_name_list)

    w = xw.Book()
    new_list = []
    for i in missing_name_list:
        new_list.append([i])
    xw.Range('A1').value = new_list 

def replace_role(filename):
    print filename
    wb = xw.Book(filename)
    replace_time = 0
 
    write_to = []
    for row in range(1, 932):
        first_name = str(wb.sheets[0].range('A'+str(row)).value).lower().encode('ascii','ignore').strip()
        last_name = str(wb.sheets[0].range('B'+str(row)).value).lower().encode('ascii','ignore').strip()
        sso = wb.sheets[0].range('C'+str(row)).value
        if sso:
            sso = str(sso.encode('ascii','ignore')).lower().strip()
        else:
            sso = ""

        bu_real = wb.sheets[0].range('D'+str(row)).value
        if bu_real:
            bu_real = str(bu_real.encode('ascii','ignore')).lower().strip()
        else:
            bu_real = ""

        bu_1 = wb.sheets[0].range('E'+str(row)).value
        if bu_1:
            bu_1 = str(bu_1).encode('ascii','ignore').lower().strip()
        else:
            bu_1 = ""

        bu_2 = wb.sheets[0].range('F'+str(row)).value
        if bu_2:
            bu_2 = str(bu_2.encode('ascii','ignore')).lower().strip()
        else:
            bu_2 = ""

        bu_3 = wb.sheets[0].range('G'+str(row)).value
        if bu_3:
            bu_3 = str(bu_3.encode('ascii','ignore')).lower().strip()
        else:
            bu_3 = ""

        combi_name = last_name+', '+first_name

        new_list = []
        new_list.append(first_name)
        new_list.append(last_name)
        new_list.append(sso)

        qe = "quality engineering"
        foundation = "foundation"

        if qe in bu_1 or qe in bu_2 or qe in bu_3:
            bu_real = "QE"
            replace_time+=1
        if foundation in bu_real:
            bu_real = "TES"

        new_list.append(bu_real)

        write_to.append(new_list)

    wb = xw.Book()
    xw.Range('A1').value = write_to 
