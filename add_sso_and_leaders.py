#!/usr/bin/python

''' This script reads a spreadsheet which has a lastname and firstname
    separated by a comma in the first column and creates another spreadsheet
    with four columns: lastname, firstname, SSOid and PCI manager
'''
from openpyxl import load_workbook, Workbook
import sys
import ldap


def bind(server, user, cred):
    l = None

    try:
        l = ldap.initialize('ldap://' + server)
        l.simple_bind_s(user, cred)
    except ldap.LDAPError, error_message:
        print 'ERROR: LDAP connection error: %s' % error_message
        quit()
    return l


def lookup(l, filter, attribs):
    base = 'ou=users,o=rackspace'
    scope = ldap.SCOPE_SUBTREE
    retrieve_attributes = None

    try:
        matchedList = l.search_s(base, scope, filter, attribs)
        if not matchedList:
            print 'No match in LDAP for the parameters given '
            return None

        for m in matchedList:
            matchedDN = m[0]
            #print 'LDAP matched=' + str(matchedDN)
            attrDict = m[1]
            return attrDict

    except ldap.LDAPError, error_message:
        print 'ERROR: LDAP lookup error: %s' % error_message
        return None

    return None


manager_names = { 'pras5455': 'Prashanth Chandrasekar',
                'larr8291': 'Larry Browder',
                'gine3970': 'Gigi Geoffrion',
                'laur4771': 'Lauren Luensmann',
                'paul.voccio': 'Paul Voccio',
                'dani4801': 'Dan Spraggins',
                'ghrncir': 'Greg Hrncir',
                'bran7973': 'Brannon Lacey',
                'darren.norfolk': 'Darren Norfolk',
                'bria9648': 'Brian Stein',
                'stephen.nolan': 'Stephen Nolan',
                'rama9976': 'Ramakant Pandrangi',
                'todd7280': 'Todd Mitchell',
                'jim.thorpe': 'Jim Thorpe',
                'jim.hawkins': 'Jim Hawkins',
                'walt.leddy': 'Walt Leddy',
                'rmoore': 'Rusty Moore',
                'brya8645': 'Bryan Law',
                'misc': 'misc',
                'error': 'error',
                'taylor.rhodes': 'misc',
               }


manager_uids = manager_names.keys()


attribs = ['givenName',
               'sn',
               'mail',
               'uid',
               'ou',
               'manager',
              ]

server = 'auth.edir.rackspace.com'


def managersManagerList(manager_uid, managers_list, bound_server):
    ''' This is a recursive function which given a SSO id, walks the LDAP tree 
        upwards and collects the hierarchy of managers and return them as a list
    '''
    if manager_uid == 'taylor.rhodes' or manager_uid == 'error':
        return managers_list

    filter = '(uid=' + manager_uid + ')'
    attrDict = lookup(bound_server, filter, attribs)
    if attrDict:
        if 'manager' in attrDict:
            manager_manager_uid = str(attrDict['manager'][0])[3:-21]
            #print 'manager manager=' + manager_manager_uid
        else:
            print 'manager key is not in attrDict for manager_uid=' + manager_uid
            manager_manager_uid = 'taylor.rhodes' #temporary
    else:
        manager_uid = 'error'
        print 'ERROR: ' + manager_uid + ' has no manager in LDAP !!!!!!'
        managers_list.append(manager_uid)
        return managers_list

    managers_list.append(manager_uid)
    return managersManagerList(manager_manager_uid, managers_list, bound_server)


def getEmployeeManager(firstname, lastname, bound_server):
    ''' This function gets the Racker's immediate manager SSO id given a Racker's name '''
    filter = '(&(sn=' + lastname + ')' + '(givenName=' + firstname + '))'
    attrDict = lookup(bound_server, filter, attribs)
    if attrDict:
        if 'manager' in attrDict:
            manager_uid = str(attrDict['manager'][0])[3:-21]
            racker_uid = attrDict['uid'][0]
        else:
            print 'manager key is not in attrDict for racker=' + firstname + ' ' + lastname
            manager_uid = 'error' # temporary
            racker_uid = ''
    else:
        manager_uid = 'error'
        racker_uid = ''
    
    return (manager_uid, racker_uid)


def returnBusinessUnitManager(rackerfn, rackerln, bound_server):
    ''' This function checks whether a Racker falls under one of the managers that
        need to undergo PCI training.  If so, it will return that manager's SSO id
    '''
    #print 'rackerfn=' + rackerfn + '     ' + 'rackerln=' + rackerln
    (manager_uid, racker_uid) = getEmployeeManager(rackerfn, rackerln, bound_server)
    #print 'manager_uid=' + manager_uid
    #print 'racker_uid=' + racker_uid
    
    managers_list = []
    if manager_uid == 'error':
        print 'ERROR: manager_uid supplied is error'
        return 'error'
    else:
        managers_list = managersManagerList(manager_uid, managers_list, bound_server)
        if len(managers_list) > 0 and managers_list[-1] == 'error':
            return 'error'
    
    print str(managers_list)
    
    # compares the two lists to see if there is an overlap between the Racker's
    # manager hierarchy and the list of managers required to have PCI training
    for mgr_uid in manager_uids:
        if mgr_uid in managers_list:
            print 'found manager for BU: ' + mgr_uid
            return mgr_uid
    else:
        print 'this user does not fall under any of the PCI managers in the managers_list'
        return 'misc'


# ---------------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------------
if len(sys.argv) > 1:
        in_spreadsheet = sys.argv[1]
        index = in_spreadsheet.rfind('.')
        if index != -1:
                if in_spreadsheet[index:] != '.xlsx' and in_spreadsheet[index:] != '.xls':
                        print 'Argument supplied is not a spreadsheet name'
                        quit()
        else:
                print 'Argument does not end with .xlsx or .xls'
                quit()
else:
        print 'Usage:  python add_sso_and_leaders.py <path to spreadsheet>'
        quit()


server = 'auth.edir.rackspace.com'
bound_server = bind(server, '', '')

in_spreadsheet = in_spreadsheet.strip()
if 'xlsx' in in_spreadsheet:
    out_spreadsheet = in_spreadsheet[0:-5] + '_with_leaders.xlsx'
else:
    out_spreadsheet = in_spreadsheet[0:-4] + '_with_leaders.xlsx'

out_wb = Workbook()
out_ws = out_wb.active

in_wb = load_workbook(in_spreadsheet)
in_ws = in_wb.active

MAX_ROWS=6000
row_count = 1
for row in in_ws.iter_rows(min_row=1, max_col=1, max_row=MAX_ROWS):
        user = row[0].value
        if not user:
            break

        name = user.split(',')
        lastname = name[0]
        firstname = name[1]
        if firstname and lastname:
            firstname = firstname.strip()
            lastname = lastname.strip()
            firstname = firstname.replace('(', '\(')
            firstname = firstname.replace(')', '\)')
            lastname = lastname.replace('(', '\(')
            lastname = lastname.replace(')', '\)')
            print 'lastname=' + lastname
            print 'firstname=' + firstname
        else:
            print 'ERROR: Either comma is missing or only one of firstname/lastname found for ' + user
            continue

        manager_uid = returnBusinessUnitManager(firstname, lastname, bound_server)

        filter = '(&(sn=' + lastname + ')' + '(givenName=' + firstname + '))'
        attrDict = lookup(bound_server, filter, attribs)
        if attrDict:
                out_ws.cell(row=row_count, column=1, value=str(attrDict['givenName'][0]))
                out_ws.cell(row=row_count, column=2, value=str(attrDict['sn'][0]))
                out_ws.cell(row=row_count, column=3, value=str(attrDict['uid'][0]))
                out_ws.cell(row=row_count, column=4, value=manager_names[manager_uid])
                #    str(attrDict['sn'][0]) + ', ' +
                #    str(attrDict['uid'][0]) + ', ' +
                #    manager_names[manager_uid] + '\n'
        else:
                out_ws.cell(row=row_count, column=1, value=firstname)
                out_ws.cell(row=row_count, column=2, value=lastname)
                out_ws.cell(row=row_count, column=3, value='unknown')
                out_ws.cell(row=row_count, column=4, value='unknown')
                #    lastname + ', ' +
                #    'unknown' + ', ' +
                #    'unknown' + '\n' 
        print '\n'
        row_count = row_count + 1

try:
    out_wb.save(out_spreadsheet)
except :
    print 'ERROR: Error saving output spreadsheet ' + out_spreadsheet

bound_server.unbind_s()

